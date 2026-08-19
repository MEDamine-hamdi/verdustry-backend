import io
from datetime import datetime

from sqlalchemy.orm import Session
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill

from app.models.company import Company
from app.models.emission import Emission
from app.models.target import Target
from app.services.benchmark_service import BenchmarkService


class ReportService:
    def __init__(self, db: Session):
        self.db = db

    def _get_data(self, company_id: int):
        company = self.db.query(Company).filter(Company.id == company_id).first()
        emissions = self.db.query(Emission).filter(Emission.company_id == company_id).all()
        targets = self.db.query(Target).filter(Target.company_id == company_id).all()
        benchmark = BenchmarkService(self.db).get_benchmark(company_id)
        return company, emissions, targets, benchmark

    def generate_pdf(self, company_id: int) -> bytes:
        company, emissions, targets, benchmark = self._get_data(company_id)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f"Rapport ESG — {company.name}", styles["Title"]))
        elements.append(Paragraph(
            f"Secteur: {company.sector or 'N/A'} | Généré le {datetime.now().strftime('%d/%m/%Y')}",
            styles["Normal"],
        ))
        elements.append(Paragraph(
            "Ce rapport suit la structure de reporting ESG conforme aux exigences CSRD, CBAM et GRI "
            "(Scope 1/2/3, cibles, écarts sectoriels).",
            styles["Normal"],
        ))
        elements.append(Spacer(1, 0.6 * cm))

        elements.append(Paragraph("Émissions par période et par scope", styles["Heading2"]))
        by_period: dict[str, dict[int, float]] = {}
        for e in emissions:
            by_period.setdefault(e.period, {}).setdefault(e.scope, 0.0)
            by_period[e.period][e.scope] += e.value

        data = [["Période", "Scope 1 (tCO2e)", "Scope 2 (tCO2e)", "Scope 3 (tCO2e)", "Total"]]
        for period, scopes in sorted(by_period.items()):
            s1, s2, s3 = scopes.get(1, 0), scopes.get(2, 0), scopes.get(3, 0)
            data.append([period, f"{s1:.2f}", f"{s2:.2f}", f"{s3:.2f}", f"{(s1+s2+s3):.2f}"])

        table = Table(data, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2f5233")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.6 * cm))

        elements.append(Paragraph("Cibles ESG", styles["Heading2"]))
        target_data = [["Nom", "Indicateur", "Référence", "Cible", "Échéance"]]
        for t in targets:
            target_data.append([
                t.name, t.metric,
                f"{t.baseline_value} ({t.baseline_year})",
                f"{t.target_value} ({t.target_year})",
                str(t.deadline) if t.deadline else "-",
            ])
        if len(target_data) > 1:
            ttable = Table(target_data, hAlign="LEFT")
            ttable.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2f5233")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]))
            elements.append(ttable)
        elements.append(Spacer(1, 0.6 * cm))

        elements.append(Paragraph("Benchmark sectoriel & réglementaire", styles["Heading2"]))
        bench_data = [["Référence", "Valeur référence", "Valeur entreprise", "Écart", "Écart %"]]
        for item in benchmark.items:
            bench_data.append([
                item.label, f"{item.referenceValue:.2f}", f"{item.companyValue:.2f}",
                f"{item.gapValue:.2f}", f"{item.gapPercent:.1f}%",
            ])
        if len(bench_data) > 1:
            btable = Table(bench_data, hAlign="LEFT")
            btable.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2f5233")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]))
            elements.append(btable)

        doc.build(elements)
        return buffer.getvalue()

    def generate_excel(self, company_id: int) -> bytes:
        company, emissions, targets, benchmark = self._get_data(company_id)
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        ws1 = wb.active
        ws1.title = "Emissions"
        ws1.append(["Période", "Site ID", "Scope", "Catégorie", "Valeur", "Unité"])
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
        for e in emissions:
            ws1.append([e.period, e.site_id, e.scope, e.category, e.value, e.unit])

        ws2 = wb.create_sheet("Cibles")
        ws2.append(["Nom", "Indicateur", "Base", "Année base", "Cible", "Année cible", "Échéance"])
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
        for t in targets:
            ws2.append([t.name, t.metric, t.baseline_value, t.baseline_year,
                        t.target_value, t.target_year, str(t.deadline) if t.deadline else ""])

        ws3 = wb.create_sheet("Benchmark")
        ws3.append(["Référence", "Type", "Valeur référence", "Valeur entreprise", "Écart", "Écart %", "Année"])
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
        for item in benchmark.items:
            ws3.append([item.label, item.referenceType, item.referenceValue,
                        item.companyValue, item.gapValue, item.gapPercent, item.year])

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()